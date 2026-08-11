"""
报告导出模块 - 将仿真结果导出为 Excel / PDF

功能：
- export_excel(): 导出多 Sheet 的 Excel 报告（KPI、工序产出、报警、WIP 采样、切换事件）
- export_pdf(): 导出 PDF 摘要报告（中文字体优先，缺失时回退 Helvetica）
- export_report(): 按文件扩展名自动分派
"""

import os
from typing import List, Optional

from src.models import SimulationResult


# 中文字体候选路径（按优先级）
_CN_FONT_CANDIDATES = [
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/System/Library/Fonts/Hiragino Sans GB.ttc",
    "/System/Library/Fonts/STHeiti Medium.ttc",
    "/System/Library/Fonts/Supplemental/Songti.ttc",
]

_CN_FONT_NAME = None
_CN_FONT_REGISTERED = False


def _register_cn_font() -> str:
    """
    注册中文字体，返回字体名称

    Returns:
        str: 可用的字体名称（找不到中文字体时回退 "Helvetica"）
    """
    global _CN_FONT_NAME, _CN_FONT_REGISTERED
    if _CN_FONT_REGISTERED:
        return _CN_FONT_NAME

    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont

        for path in _CN_FONT_CANDIDATES:
            if not os.path.exists(path):
                continue
            try:
                font_name = f"CN_{os.path.basename(path).replace('.', '_')}"
                pdfmetrics.registerFont(TTFont(font_name, path))
                _CN_FONT_NAME = font_name
                break
            except Exception:
                continue
    except Exception:
        pass

    if _CN_FONT_NAME is None:
        _CN_FONT_NAME = "Helvetica"
    _CN_FONT_REGISTERED = True
    return _CN_FONT_NAME


def _kpi_rows(result: SimulationResult) -> List[List[str]]:
    """构造 KPI 表格行（中文标签 + 值）"""
    k = result.kpis
    duration_minutes = result.duration_seconds / 60.0
    unit = result.unit
    return [
        ["产线名称", result.line_name],
        ["仿真时长(分钟)", f"{duration_minutes:.1f}"],
        [f"总产出({unit})", f"{result.total_output}"],
        [f"可发运产出({unit})", f"{k.get('shippable_quantity', result.total_output)}"],
        ["隔离批次", f"{k.get('rejected_batches', 0)}"],
        [f"瓶颈产能({unit}/h)", f"{k.get('bottleneck_capacity', 0):.1f}"],
        [f"预计日产量({unit})", f"{k.get('daily_output', 0):.0f}"],
        ["日成本(元)", f"{k.get('total_cost', 0):.0f}"],
        [f"单位成本(元/{unit})", f"{k.get('unit_cost', 0):.3f}"],
        ["产线平衡率", f"{k.get('balance_rate', 0) * 100:.1f}%"],
        [f"UPPH({unit}/人·h)", f"{k.get('upph', 0):.1f}"],
    ]


def export_excel(result: SimulationResult, file_path: str) -> bool:
    """
    导出 Excel 报告

    Args:
        result: 仿真结果
        file_path: 保存路径（.xlsx）

    Returns:
        bool: 是否成功
    """
    try:
        import pandas as pd

        kpi_df = pd.DataFrame(_kpi_rows(result), columns=["指标", "值"])

        station_rows = []
        for station in result.station_outputs:
            station_rows.append({
                "工序ID": station,
                f"产出({result.unit})": result.station_outputs.get(station, 0),
                f"WIP({result.unit})": result.station_wips.get(station, 0),
            })
        station_df = pd.DataFrame(station_rows)

        alert_df = pd.DataFrame([
            {
                "时间(分钟)": round(a.timestamp_minutes, 1),
                "级别": a.severity,
                "类型": a.alert_type,
                "消息": a.message,
                "建议": a.suggestion,
            }
            for a in result.alerts
        ])

        wip_df = pd.DataFrame(result.wip_samples)
        changeover_df = pd.DataFrame(result.changeover_events)
        batch_df = pd.DataFrame(result.batch_results)
        quality_df = pd.DataFrame(result.quality_results)
        metrics_df = pd.DataFrame([
            {
                '工序ID': sid,
                '工序名': m.get('name', ''),
                '运行秒': m.get('running_sec', 0),
                '等待秒': m.get('waiting_sec', 0),
                '堵塞秒': m.get('blocked_sec', 0),
                '实际利用率': m.get('utilization', 0),
                'OEE可用率': m.get('oee_availability', 0),
                'OEE性能率': m.get('oee_performance', 0),
                'OEE合格率': m.get('oee_quality', 0),
                'OEE': m.get('oee_total', 0),
            }
            for sid, m in result.station_metrics.items()
        ])

        # P2 失衡分析
        imbalance_rows = []
        station_ids = list(result.station_metrics.keys())
        for i, sid in enumerate(station_ids):
            m = result.station_metrics[sid]
            ratio = None
            if i > 0:
                prev_cap = result.station_metrics[station_ids[i - 1]].get('capacity', 0)
                capacity = m.get('capacity', 0)
                ratio = round(capacity / prev_cap, 2) if prev_cap else None
            conclusion = []
            if m.get('utilization', 0) >= 0.95:
                conclusion.append('瓶颈')
            if m.get('waiting_sec', 0) >= 300:
                conclusion.append('饥饿')
            if m.get('blocked_sec', 0) >= 300:
                conclusion.append('堵塞')
            if ratio is not None and ratio < 0.8:
                conclusion.append('产能冗余')
            imbalance_rows.append({
                '工序ID': sid,
                '工序名': m.get('name', ''),
                '理论产能': m.get('capacity', 0),
                '上下游产能比': ratio if ratio is not None else '',
                '实际利用率': m.get('utilization', 0),
                '等待秒': m.get('waiting_sec', 0),
                '堵塞秒': m.get('blocked_sec', 0),
                '结论': '/'.join(conclusion) or '正常',
            })
        imbalance_df = pd.DataFrame(imbalance_rows)

        # V3.2 原料事件
        material_df = pd.DataFrame(result.material_events)

        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            kpi_df.to_excel(writer, sheet_name="KPI", index=False)
            station_df.to_excel(writer, sheet_name="工序产出", index=False)
            alert_df.to_excel(writer, sheet_name="报警记录", index=False)
            wip_df.to_excel(writer, sheet_name="WIP采样", index=False)
            changeover_df.to_excel(writer, sheet_name="切换事件", index=False)
            if not batch_df.empty:
                batch_df.to_excel(writer, sheet_name="批次结果", index=False)
            if not quality_df.empty:
                quality_df.to_excel(writer, sheet_name="质量门", index=False)
            if not metrics_df.empty:
                metrics_df.to_excel(writer, sheet_name="工序指标", index=False)
            if not imbalance_df.empty:
                imbalance_df.to_excel(writer, sheet_name="失衡分析", index=False)
            if not material_df.empty:
                material_df.to_excel(writer, sheet_name="原料事件", index=False)

        return True
    except Exception:
        return False


def export_pdf(result: SimulationResult, file_path: str) -> bool:
    """
    导出 PDF 摘要报告

    Args:
        result: 仿真结果
        file_path: 保存路径（.pdf）

    Returns:
        bool: 是否成功
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        font_name = _register_cn_font()

        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        doc = SimpleDocTemplate(
            file_path,
            pagesize=A4,
            rightMargin=20 * mm,
            leftMargin=20 * mm,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CNTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=18,
            leading=24,
        )
        body_style = ParagraphStyle(
            "CNBody",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=10,
            leading=14,
        )
        header_style = ParagraphStyle(
            "CNHeader",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=11,
            leading=14,
        )

        story = [
            Paragraph(f"电子烟产线仿真报告 - {result.line_name}", title_style),
            Spacer(1, 8 * mm),
            Paragraph(
                f"仿真时长：{result.duration_seconds / 60:.1f} 分钟　"
                f"总产出：{result.total_output} {result.unit}",
                body_style,
            ),
            Spacer(1, 6 * mm),
            Paragraph("一、核心 KPI", header_style),
            Spacer(1, 3 * mm),
        ]

        kpi_table = Table(_kpi_rows(result), colWidths=[80 * mm, 90 * mm])
        kpi_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (0, -1), colors.lightgrey),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(kpi_table)

        if result.alerts:
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph("二、报警记录", header_style))
            story.append(Spacer(1, 3 * mm))
            alert_rows = [["时间(分钟)", "级别", "消息"]]
            for alert in result.alerts[:50]:
                alert_rows.append([
                    f"{alert.timestamp_minutes:.1f}",
                    alert.severity,
                    alert.message,
                ])
            alert_table = Table(alert_rows, colWidths=[25 * mm, 25 * mm, 120 * mm])
            alert_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(alert_table)

        if result.batch_results:
            story.append(Spacer(1, 6 * mm))
            story.append(Paragraph("三、批次结果", header_style))
            story.append(Spacer(1, 3 * mm))
            batch_rows = [["批次ID", "配方", "周期(分钟)", "合格率", "产出(L)"]]
            for b in result.batch_results:
                batch_rows.append([
                    str(b.get("batch_id", "")),
                    str(b.get("recipe_name", "")),
                    f"{b.get('cycle_min', 0):.1f}",
                    f"{b.get('pass_rate', 0) * 100:.1f}%",
                    f"{b.get('yield_l', 0):.1f}",
                ])
            batch_table = Table(batch_rows, colWidths=[30 * mm, 35 * mm, 35 * mm, 30 * mm, 40 * mm])
            batch_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(batch_table)

        story.append(Spacer(1, 6 * mm))
        story.append(Paragraph(
            "注：本报告由仿真工具自动生成，结果作为产线设计参考，"
            "建议结合实际生产数据校准 OEE 等参数。",
            body_style,
        ))

        doc.build(story)
        return True
    except Exception:
        return False


def export_hr_report(summary: dict, file_path: str) -> bool:
    """导出人力规划 Excel 报告（V3.3 迭代）"""
    try:
        import pandas as pd
        costs = summary.get("costs", {})

        overview_df = pd.DataFrame([
            ["目标日产量", summary.get("daily_target", 0)],
            ["每日有效工时(小时)", summary.get("effective_hours_per_day", 0)],
            ["每小时需求", summary.get("required_hourly", 0)],
            ["人力需求合计(最低人数)", summary.get("total_headcount", 0)],
            ["当前在岗合计", summary.get("current_total", 0)],
            ["富余人数(在岗−需求)", summary.get("excess_headcount", 0)],
            ["一次性招聘建议", summary.get("one_time_hiring", 0)],
            ["首周净缺口(需求−在岗−到岗)", summary.get("initial_gap", 0)],
            ["达产天数", summary.get("days_to_full", 0)],
        ], columns=["指标", "值"])

        derivation_df = pd.DataFrame([
            {
                "工序": d.get("name", ""),
                "是否瓶颈": "是" if d.get("is_bottleneck") else "",
                "配置人数": d.get("configured", 0),
                "单件耗时(秒)": d.get("process_time", 0),
                "机台节拍(秒)": d.get("machine_takt") or "",
                "OEE": d.get("oee", 0),
                "效率": d.get("efficiency", 0),
                "切换+清洗(分钟)": d.get("changeover_minutes", 0),
                "协作模式": d.get("collaboration", ""),
                "工序产能(/h)": d.get("capacity", 0),
                "单人产能(/h)": d.get("per_head", 0),
                "需求人数": d.get("needed", 0),
            }
            for d in summary.get("station_derivation", [])
        ])

        role_df = pd.DataFrame([
            {"工种": role, "需求人数": n}
            for role, n in summary.get("headcount_by_role", {}).items()
        ] + [{
            "工种": "合计", "需求人数": summary.get("total_headcount", 0),
        }])

        current_df = pd.DataFrame([
            {"工种": role, "在岗人数": n}
            for role, n in summary.get("current_headcount", {}).items()
        ] + [{
            "工种": "合计", "在岗人数": summary.get("current_total", 0),
        }])

        cost_df = pd.DataFrame([
            ["总人数", costs.get("headcount", 0)],
            ["在岗覆盖人数（含缺勤）", costs.get("covered_headcount", 0)],
            ["日人力成本(元)", costs.get("daily_labor_cost", 0)],
            ["月人力成本(元)", costs.get("monthly_labor_cost", 0)],
            ["月招聘/培训(元)", costs.get("monthly_recruit_training", 0)],
            ["月总成本(元)", costs.get("monthly_total", 0)],
            ["单位人力成本(元/单位)", costs.get("per_unit_labor_cost", 0)],
        ], columns=["指标", "值"])

        gap_rows = []
        for row in summary.get("weekly_gap", []):
            gap_rows.append({
                "周": row.get("week", 0),
                "周底还缺(累计)": row.get("total_gap", 0),
                "本周新增需求": row.get("new_gap", 0),
                **{f"缺口_{role}": v for role, v in row.get("gap", {}).items()},
            })
        gap_df = pd.DataFrame(gap_rows)

        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            overview_df.to_excel(writer, sheet_name="概览", index=False)
            derivation_df.to_excel(writer, sheet_name="计算明细", index=False)
            role_df.to_excel(writer, sheet_name="人力需求", index=False)
            current_df.to_excel(writer, sheet_name="当前在岗", index=False)
            cost_df.to_excel(writer, sheet_name="成本", index=False)
            gap_df.to_excel(writer, sheet_name="招聘缺口", index=False)

        from openpyxl import load_workbook
        from openpyxl.styles import Font

        workbook = load_workbook(file_path)
        header_font = Font(bold=True, color="FFFFFF")
        for sheet_name, widths in (
            ("概览", [36, 20]),
            ("计算明细", [16, 9, 9, 12, 12, 8, 8, 13, 10, 12, 12, 9]),
            ("人力需求", [24, 12]),
            ("当前在岗", [24, 12]),
            ("成本", [30, 18]),
            ("招聘缺口", [10, 14, 14]),
        ):
            worksheet = workbook[sheet_name]
            for cell in worksheet[1]:
                cell.font = header_font
            for idx, width in enumerate(widths, 1):
                worksheet.column_dimensions[
                    worksheet.cell(row=1, column=idx).column_letter
                ].width = width
            worksheet.freeze_panes = "A2"
        workbook.save(file_path)
        return True
    except Exception:
        return False


def export_hr_pdf(summary: dict, file_path: str) -> bool:
    """导出人力规划 PDF 报告（V3.3 迭代）"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        font_name = _register_cn_font()
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)

        doc = SimpleDocTemplate(
            file_path, pagesize=A4,
            rightMargin=20 * mm, leftMargin=20 * mm,
            topMargin=20 * mm, bottomMargin=20 * mm,
        )
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CNTitle", parent=styles["Title"],
            fontName=font_name, fontSize=18, leading=24,
        )
        body_style = ParagraphStyle(
            "CNBody", parent=styles["Normal"],
            fontName=font_name, fontSize=10, leading=14,
        )

        costs = summary.get("costs", {})

        def _table(rows, widths, header=True):
            table = Table(rows, colWidths=widths, repeatRows=1 if header else 0)
            style = [
                ("FONTNAME", (0, 0), (-1, -1), font_name),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
            if header:
                style += [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2329")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ]
            table.setStyle(TableStyle(style))
            return table

        story = [
            Paragraph("人力规划报告", title_style),
            Spacer(1, 6 * mm),
            Paragraph("一、概览", body_style),
            Spacer(1, 2 * mm),
        ]
        overview_rows = [
            ["指标", "值"],
            ["目标日产量", summary.get("daily_target", 0)],
            ["每日有效工时(小时)", summary.get("effective_hours_per_day", 0)],
            ["每小时需求", summary.get("required_hourly", 0)],
            ["人力需求合计(最低人数)", summary.get("total_headcount", 0)],
            ["当前在岗合计", summary.get("current_total", 0)],
            ["富余人数(在岗−需求)", summary.get("excess_headcount", 0)],
            ["一次性招聘建议", summary.get("one_time_hiring", 0)],
            ["达产天数", summary.get("days_to_full", 0)],
        ]
        story.append(_table(overview_rows, [110 * mm, 50 * mm]))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("二、计算明细", body_style))
        story.append(Spacer(1, 2 * mm))
        derivation_rows = [["工序", "配置", "单人产能", "需求", "备注"]]
        derivation_rows += [
            [
                d.get("name", ""),
                str(d.get("configured", 0)),
                f"{d.get('per_head', 0)}/h",
                str(d.get("needed", 0)),
                "瓶颈" if d.get("is_bottleneck") else "",
            ]
            for d in summary.get("station_derivation", [])
        ]
        story.append(_table(
            derivation_rows, [60 * mm, 20 * mm, 45 * mm, 20 * mm, 30 * mm]
        ))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("三、人力需求（按工种）", body_style))
        story.append(Spacer(1, 2 * mm))
        role_rows = [["工种", "所需人数"]]
        role_rows += [
            [role, str(count)]
            for role, count in summary.get("headcount_by_role", {}).items()
        ]
        role_rows.append(["合计", str(summary.get("total_headcount", 0))])
        story.append(_table(role_rows, [80 * mm, 40 * mm]))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("四、成本测算", body_style))
        story.append(Spacer(1, 2 * mm))
        cost_rows = [
            ["指标", "值"],
            ["总人数", costs.get("headcount", 0)],
            ["在岗覆盖人数（含缺勤）", costs.get("covered_headcount", 0)],
            ["日人力成本(元)", costs.get("daily_labor_cost", 0)],
            ["月总成本(元)", costs.get("monthly_total", 0)],
            ["单位人力成本(元/单位)", costs.get("per_unit_labor_cost", 0)],
        ]
        story.append(_table(cost_rows, [80 * mm, 40 * mm]))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph("五、招聘缺口（前 12 周）", body_style))
        story.append(Spacer(1, 2 * mm))
        gap_rows = [["周", "周底还缺(累计)", "本周新增需求"]]
        gap_rows += [
            [
                str(row.get("week", 0)),
                str(row.get("total_gap", 0)),
                str(row.get("new_gap", 0)),
            ]
            for row in summary.get("weekly_gap", [])[:12]
        ]
        story.append(_table(gap_rows, [30 * mm, 70 * mm, 70 * mm]))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(
            f"六、结论：一次性招聘建议 {summary.get('one_time_hiring', 0)} 人；"
            f"预计 {summary.get('days_to_full', 0)} 天达产。",
            body_style,
        ))
        doc.build(story)
        return True
    except Exception:
        return False


def export_report(result: SimulationResult, file_path: str) -> bool:
    """
    按文件扩展名导出报告

    Args:
        result: 仿真结果
        file_path: 保存路径（.xlsx 或 .pdf）

    Returns:
        bool: 是否成功
    """
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".xlsx":
        return export_excel(result, file_path)
    if ext == ".pdf":
        return export_pdf(result, file_path)
    return False
