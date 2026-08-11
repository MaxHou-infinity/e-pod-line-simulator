"""报告导出单元测试"""

import os

from src.models import ProductionLine, Station, create_liquid_line
from src.reporting import (
    export_excel,
    export_hr_pdf,
    export_hr_report,
    export_pdf,
    export_report,
)
from src.simulation import SimulationEngine


def make_result():
    line = ProductionLine("报告测试", shift_hours=1, break_minutes=0)
    line.add_station(Station("s01", "注油", 1.0, 1))
    line.add_station(Station("s02", "包装", 1.0, 1))
    return SimulationEngine(line).run_sync(duration_hours=0.1)


def test_export_excel(tmp_path):
    path = str(tmp_path / "report.xlsx")
    assert export_excel(make_result(), path) is True
    assert os.path.exists(path)

    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "KPI" in wb.sheetnames
    assert "工序产出" in wb.sheetnames
    assert "报警记录" in wb.sheetnames
    assert "WIP采样" in wb.sheetnames
    assert "切换事件" in wb.sheetnames
    assert "工序指标" in wb.sheetnames
    assert "失衡分析" in wb.sheetnames


def test_export_pdf(tmp_path):
    path = str(tmp_path / "report.pdf")
    assert export_pdf(make_result(), path) is True
    assert os.path.exists(path)
    with open(path, "rb") as f:
        assert f.read(4) == b"%PDF"


def test_export_report_dispatch(tmp_path):
    xlsx_path = str(tmp_path / "a.xlsx")
    pdf_path = str(tmp_path / "b.pdf")
    bad_path = str(tmp_path / "c.txt")
    result = make_result()
    assert export_report(result, xlsx_path) is True
    assert export_report(result, pdf_path) is True
    assert export_report(result, bad_path) is False


def test_liquid_report_contains_batch_sheets(tmp_path):
    line = create_liquid_line()
    result = SimulationEngine(line).run_sync(duration_hours=24.0)
    path = str(tmp_path / "liquid_report.xlsx")
    assert export_excel(result, path) is True

    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "批次结果" in wb.sheetnames
    assert "质量门" in wb.sheetnames
    kpi_values = [
        str(c.value) for row in wb["KPI"].iter_rows() for c in row if c.value
    ]
    assert any("千克" in v for v in kpi_values)


def test_export_hr_report_excel_and_pdf(tmp_path):
    from src.hr_planning import (
        LaborCostConfig,
        LearningCurveConfig,
        ShiftPlan,
        build_hr_summary,
    )

    line = ProductionLine("HR报告测试", shift_hours=8, break_minutes=60)
    line.add_station(Station(
        "s01", "组装", 60.0, 2,
        oee=1.0, efficiency=1.0, changeover_time=0,
    ))
    summary = build_hr_summary(
        line,
        840,
        ShiftPlan(shifts_per_day=1, shift_hours=8, break_minutes=60),
        LaborCostConfig(),
        LearningCurveConfig(ramp_days=30),
        current={"general": 2},
        hiring_plan=[],
    )

    xlsx = str(tmp_path / "hr.xlsx")
    pdf = str(tmp_path / "hr.pdf")
    assert export_hr_report(summary, xlsx) is True
    assert export_hr_pdf(summary, pdf) is True

    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    assert "概览" in wb.sheetnames
    assert "计算明细" in wb.sheetnames
    assert "人力需求" in wb.sheetnames
    assert "当前在岗" in wb.sheetnames
    assert "成本" in wb.sheetnames
    assert "招聘缺口" in wb.sheetnames
    overview_values = [
        str(c.value) for row in wb["概览"].iter_rows() for c in row if c.value
    ]
    assert any("每小时需求" in v for v in overview_values)
