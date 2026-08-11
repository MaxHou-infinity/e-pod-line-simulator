"""KPI 历史趋势模块单元测试（V3.2 P1）"""

from src.history import (
    append_snapshot,
    build_snapshot,
    clear_history,
    export_history,
    kpi_series,
    load_history,
)
from src.models import ProductionLine, Station


def make_line():
    line = ProductionLine("历史测试", shift_hours=1, break_minutes=0)
    line.add_station(Station("s01", "注油", 1.0, 1))
    return line


def test_append_and_load_round_trip(tmp_path):
    path = str(tmp_path / "history.json")
    line = make_line()
    record = build_snapshot(line)

    assert append_snapshot(record, path) is True
    assert append_snapshot(record, path) is True

    history = load_history(path)
    assert len(history) == 2
    assert history[0]["line_name"] == "历史测试"
    assert "daily_output" in history[0]["kpis"]


def test_kpi_series_extracts_values(tmp_path):
    path = str(tmp_path / "history.json")
    line = make_line()
    for _ in range(3):
        append_snapshot(build_snapshot(line), path)

    series = kpi_series(load_history(path), "daily_output")
    assert len(series) == 3
    assert all(isinstance(v, float) for _, v in series)


def test_history_capped_at_max(tmp_path):
    path = str(tmp_path / "history.json")
    line = make_line()
    from src.history import MAX_RECORDS

    for _ in range(MAX_RECORDS + 10):
        append_snapshot(build_snapshot(line), path)
    assert len(load_history(path)) == MAX_RECORDS


def test_export_history_excel(tmp_path):
    path = str(tmp_path / "history.json")
    line = make_line()
    append_snapshot(build_snapshot(line), path)
    out = str(tmp_path / "history.xlsx")

    assert export_history(out, load_history(path)) is True

    from openpyxl import load_workbook

    wb = load_workbook(out)
    assert "Sheet1" in wb.sheetnames


def test_clear_history_removes_file(tmp_path):
    path = str(tmp_path / "history.json")
    line = make_line()
    append_snapshot(build_snapshot(line), path)

    assert clear_history(path) is True
    assert load_history(path) == []
